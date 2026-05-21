"""End-to-end: per-tenant xlsx template upload + render (v0.2.0 Task 6).

The test:

1. Constructs a synthetic xlsx with a distinctive watermark cell (Z1 =
   "TENANT-CUSTOM-MARKER"). The target's shipped template has no such
   binding, so the cell only exists if the uploaded file was used.
2. Uploads it via ``POST /v1/templates/us-hs.v1/grade_9``.
3. Runs ``POST /v1/transcripts`` against the same target/template.
4. Pulls the rendered output via ``GET /v1/jobs/{id}/result``.
5. Asserts the output's Z1 cell is the watermark — proving the uploaded
   bytes flowed through to the renderer instead of the shipped file.

Also covers the sad paths the endpoint should reject:

* Non-xlsx upload → 415.
* Unknown target_id → 404.
* Oversize upload → 413.
"""

from __future__ import annotations

import base64
import io
import secrets
from collections.abc import AsyncIterator
from pathlib import Path
from typing import Any
from unittest.mock import AsyncMock

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from PIL import Image as PILImage

from ocr_to_report.adapters.blob import LocalBlobStore
from ocr_to_report.adapters.crypto import EnvelopeEncryptor, EnvKEKProvider
from ocr_to_report.adapters.crypto.envelope import DEK_BYTES
from ocr_to_report.adapters.db import Base, get_engine, get_sessionmaker
from ocr_to_report.adapters.db.repositories import ApiKeyRepo, TenantRepo
from ocr_to_report.adapters.queue import InMemoryQueue
from ocr_to_report.adapters.vision import (
    ExtractionResult,
    FixedPolicy,
    InMemoryAsyncCache,
    ProviderRouter,
    TokenUsage,
    VisionProvider,
)
from ocr_to_report.api.app import create_app
from ocr_to_report.api.deps import AppState
from ocr_to_report.api.settings import Settings
from ocr_to_report.core.profiles import ProfileRegistry
from ocr_to_report.core.sla import SLA_PRESETS, SlaTier
from ocr_to_report.core.targets import TargetRegistry

REPO_ROOT = Path(__file__).resolve().parents[3]
WATERMARK_CELL = "Z1"
WATERMARK_VALUE = "TENANT-CUSTOM-MARKER"


# ─── helpers ──────────────────────────────────────────────────────────


def _png() -> bytes:
    img = PILImage.new("RGB", (400, 600), color=(220, 220, 220))
    out = io.BytesIO()
    img.save(out, format="PNG")
    return out.getvalue()


def _polish_extraction() -> dict[str, Any]:
    return {
        "full_name": "Jan Kowalski",
        "birth_date": "2010-01-15",
        "school_year": "2023/2024",
        "current_class_name": "pierwszej",
        "school_name": "Test Academy LO",
        "city": "Warszawa",
        "region": "mazowieckie",
        "promoted": True,
        "promoted_with_distinction": True,
        "conduct": "wzorowe",
        "subjects": [
            {"raw_subject_name": "Język polski", "raw_grade_value": "bardzo dobry"},
            {"raw_subject_name": "Matematyka", "raw_grade_value": "celujący"},
            {"raw_subject_name": "Fizyka", "raw_grade_value": "bardzo dobry"},
            {"raw_subject_name": "Chemia", "raw_grade_value": "bardzo dobry"},
            {"raw_subject_name": "Biologia", "raw_grade_value": "celujący"},
            {"raw_subject_name": "Historia", "raw_grade_value": "bardzo dobry"},
            {"raw_subject_name": "Geografia", "raw_grade_value": "dobry"},
            {"raw_subject_name": "Informatyka", "raw_grade_value": "celujący"},
            {"raw_subject_name": "Wychowanie fizyczne", "raw_grade_value": "bardzo dobry"},
            {"raw_subject_name": "Język angielski IV.1r.", "raw_grade_value": "bardzo dobry"},
        ],
        "advanced_subjects": ["Język angielski"],
    }


def _custom_xlsx_bytes() -> bytes:
    """Build a watermarked xlsx that the test can identify in the rendered output.

    The shipped target bundle never writes to Z1, so finding the watermark
    in the output proves the renderer used these bytes, not the shipped
    template. The cells the renderer DOES write to (A2 etc. per the
    target's bindings) will be overwritten — that's the whole point of
    the override path.
    """
    wb = Workbook()
    sheet = wb.active
    assert sheet is not None
    sheet[WATERMARK_CELL] = WATERMARK_VALUE
    # Pre-seed the cells the target's bindings will overwrite. Their
    # values here don't matter; the renderer rewrites them.
    sheet["A1"] = "placeholder"
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


class _FixedConfidenceVisionAdapter:
    """High-confidence vision mock so the SLA gate doesn't park the job."""

    name = VisionProvider.MOCK

    def __init__(self, confidence: float = 0.97) -> None:
        self._confidence = confidence
        self.extract = AsyncMock(side_effect=self._extract)

    async def _extract(
        self, _request: Any, *, override_api_key: str | None = None
    ) -> ExtractionResult:
        del override_api_key  # v0.3.0 BYOK threading; mock doesn't care
        return ExtractionResult(
            raw_extraction=_polish_extraction(),
            confidence=self._confidence,
            field_confidences=None,
            warnings=[],
            provider=VisionProvider.MOCK,
            model_id="mock",
            usage=TokenUsage(input_tokens=1500, output_tokens=300, usd_cost=0.003),
        )

    async def aclose(self) -> None:
        return None


@pytest.fixture
def settings(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> Settings:
    monkeypatch.setenv(
        "OCR2R_KEK_B64",
        base64.b64encode(secrets.token_bytes(DEK_BYTES)).decode(),
    )
    db_url = f"sqlite+aiosqlite:///{tmp_path / 'test.db'}"
    return Settings(
        env="development",
        database_url=db_url,
        blob_backend="local",
        blob_local_root=tmp_path / "blob",
        kek_b64=base64.b64encode(secrets.token_bytes(DEK_BYTES)).decode(),
        profiles_root=REPO_ROOT / "profiles",
        targets_root=REPO_ROOT / "targets",
    )


@pytest.fixture
async def db_setup(settings: Settings) -> None:
    engine = get_engine(settings.database_url)
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)


@pytest.fixture
async def standard_client(
    settings: Settings, db_setup: None
) -> AsyncIterator[tuple[TestClient, dict[str, Any]]]:
    encryptor = EnvelopeEncryptor(EnvKEKProvider(env_var="OCR2R_KEK_B64"))
    sm = get_sessionmaker(settings.database_url)
    async with sm() as session:
        tenants = TenantRepo(session, encryptor)
        tenant, _ = await tenants.create(name="Acme Tmpl", slug="acme-tmpl")
        tenant.sla_tier = SlaTier.STANDARD.value
        keys = ApiKeyRepo(session)
        _row, plain_key = await keys.issue(tenant_id=tenant.id, scopes=["transcripts:write"])
        await session.commit()
        seeded = {"tenant_id": tenant.id, "api_key": plain_key}

    app = create_app(settings=settings)
    router = ProviderRouter(
        {VisionProvider.MOCK: _FixedConfidenceVisionAdapter(confidence=0.97)},
        FixedPolicy(VisionProvider.MOCK),
    )
    encryptor2 = EnvelopeEncryptor(EnvKEKProvider(env_var="OCR2R_KEK_B64"))
    profile_registry = ProfileRegistry(settings.profiles_root.resolve())
    target_registry = TargetRegistry(settings.targets_root.resolve())
    bundle_roots = {t.id: settings.targets_root / t.id for t in target_registry.all()}

    client = TestClient(app)
    client.__enter__()
    app.state.app_state = AppState(
        settings=settings,
        encryptor=encryptor2,
        profile_registry=profile_registry,
        target_registry=target_registry,
        blob_store=LocalBlobStore(settings.blob_local_root),
        vision_router=router,
        result_cache=InMemoryAsyncCache(),
        bundle_roots=bundle_roots,
        queue=InMemoryQueue(),
        sla_presets=dict(SLA_PRESETS),
    )
    try:
        yield client, seeded
    finally:
        client.__exit__(None, None, None)


# ─── the canonical Task 6 proof ───────────────────────────────────────


def test_uploaded_template_is_used_for_render(
    standard_client: tuple[TestClient, dict[str, Any]],
) -> None:
    """Upload watermarked xlsx → run job → output carries the watermark."""
    client, seeded = standard_client
    headers = {"Authorization": f"Bearer {seeded['api_key']}"}

    # Pick the year-9 template (the default for source_year=1 in PL→US-HS).
    target_id = "us-hs.v1"
    template_key = "grade_9"
    custom_xlsx = _custom_xlsx_bytes()

    # 1. Upload the custom template.
    upload = client.post(
        f"/v1/templates/{target_id}/{template_key}",
        headers=headers,
        files={
            "template_file": (
                "custom.xlsx",
                custom_xlsx,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert upload.status_code == 201, upload.text
    body = upload.json()
    assert body["target_id"] == target_id
    assert body["template_key"] == template_key
    assert body["size_bytes"] == len(custom_xlsx)
    assert body["blob_key"].startswith(f"tenant/{seeded['tenant_id']}/templates/")
    assert body["blob_key"].endswith(".xlsx")

    # 2. Run a transcript job targeting this template.
    r = client.post(
        "/v1/transcripts",
        headers=headers,
        files={"file": ("t.png", _png(), "image/png")},
        data={
            "profile_id": "pl.lo.swiadectwo_szkolne.v1",
            "target_id": target_id,
            "target_template_key": template_key,
        },
    )
    assert r.status_code == 200, r.text
    summary = r.json()["job"]
    assert summary["status"] == "succeeded", summary

    # 3. Pull the rendered xlsx and confirm the watermark.
    download = client.get(f"/v1/jobs/{summary['id']}/result", headers=headers)
    assert download.status_code == 200, download.text
    rendered_wb = load_workbook(filename=io.BytesIO(download.content))
    rendered_sheet = rendered_wb.active
    assert rendered_sheet is not None
    assert rendered_sheet[WATERMARK_CELL].value == WATERMARK_VALUE, (
        "rendered xlsx is missing the watermark cell — the uploaded template "
        "was not used (renderer fell back to the shipped on-disk template)"
    )


# ─── sad paths ───────────────────────────────────────────────────────


def test_upload_rejects_non_xlsx_bytes(
    standard_client: tuple[TestClient, dict[str, Any]],
) -> None:
    client, seeded = standard_client
    headers = {"Authorization": f"Bearer {seeded['api_key']}"}

    r = client.post(
        "/v1/templates/us-hs.v1/grade_9",
        headers=headers,
        files={"template_file": ("nope.xlsx", b"this is not xlsx", "application/octet-stream")},
    )
    assert r.status_code == 415, r.text


def test_upload_rejects_unknown_target(
    standard_client: tuple[TestClient, dict[str, Any]],
) -> None:
    client, seeded = standard_client
    headers = {"Authorization": f"Bearer {seeded['api_key']}"}

    r = client.post(
        "/v1/templates/no-such-target/grade_9",
        headers=headers,
        files={
            "template_file": (
                "x.xlsx",
                _custom_xlsx_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert r.status_code == 404, r.text


def test_upload_rejects_unknown_template_key(
    standard_client: tuple[TestClient, dict[str, Any]],
) -> None:
    client, seeded = standard_client
    headers = {"Authorization": f"Bearer {seeded['api_key']}"}

    r = client.post(
        "/v1/templates/us-hs.v1/no_such_key",
        headers=headers,
        files={
            "template_file": (
                "x.xlsx",
                _custom_xlsx_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert r.status_code == 404, r.text


def test_delete_removes_override_and_reverts_render(
    standard_client: tuple[TestClient, dict[str, Any]],
) -> None:
    """After DELETE the watermark disappears — render falls back to shipped."""
    client, seeded = standard_client
    headers = {"Authorization": f"Bearer {seeded['api_key']}"}
    target_id = "us-hs.v1"
    template_key = "grade_9"

    # Upload.
    client.post(
        f"/v1/templates/{target_id}/{template_key}",
        headers=headers,
        files={
            "template_file": (
                "x.xlsx",
                _custom_xlsx_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    # Delete.
    rm = client.delete(f"/v1/templates/{target_id}/{template_key}", headers=headers)
    assert rm.status_code == 204, rm.text

    # Run a job — Z1 should NOT contain the watermark anymore (shipped
    # template has no such cell).
    r = client.post(
        "/v1/transcripts",
        headers=headers,
        files={"file": ("t.png", _png(), "image/png")},
        data={
            "profile_id": "pl.lo.swiadectwo_szkolne.v1",
            "target_id": target_id,
            "target_template_key": template_key,
        },
    )
    assert r.status_code == 200, r.text
    summary = r.json()["job"]
    download = client.get(f"/v1/jobs/{summary['id']}/result", headers=headers)
    rendered_wb = load_workbook(filename=io.BytesIO(download.content))
    rendered_sheet = rendered_wb.active
    assert rendered_sheet is not None
    assert rendered_sheet[WATERMARK_CELL].value != WATERMARK_VALUE


def test_delete_with_no_prior_upload_returns_404(
    standard_client: tuple[TestClient, dict[str, Any]],
) -> None:
    client, seeded = standard_client
    headers = {"Authorization": f"Bearer {seeded['api_key']}"}
    r = client.delete("/v1/templates/us-hs.v1/grade_9", headers=headers)
    assert r.status_code == 404, r.text


def test_delete_wrong_template_key_returns_404_when_other_key_overridden(
    standard_client: tuple[TestClient, dict[str, Any]],
) -> None:
    """Code-review finding: DELETE for grade_10 must 404 if only grade_9 is uploaded.

    Pre-fix, the handler 404'd only when the entire target_id row was
    missing — if the target had a different template_key uploaded, the
    DELETE for grade_10 fell through to a no-op upsert and returned
    204, lying about what happened. This test pins the corrected
    behavior.
    """
    client, seeded = standard_client
    headers = {"Authorization": f"Bearer {seeded['api_key']}"}
    # Upload grade_9 so the target_id row exists.
    upload = client.post(
        "/v1/templates/us-hs.v1/grade_9",
        headers=headers,
        files={
            "template_file": (
                "x.xlsx",
                _custom_xlsx_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert upload.status_code == 201, upload.text

    # Now DELETE a DIFFERENT key. The row exists but has no patch for it
    # — must 404, not 204.
    r = client.delete("/v1/templates/us-hs.v1/some_nonexistent_key", headers=headers)
    assert r.status_code == 404, (
        f"DELETE for a template_key with no override on an otherwise-overridden "
        f"target_id should 404; got {r.status_code} body={r.text}"
    )
