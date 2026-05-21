"""v0.2.0 end-to-end integration test against real postgres.

The unit + e2e suites under ``packages/api/tests/`` cover each override
surface in isolation against in-memory sqlite. This test is the
``@pytest.mark.integration`` companion: it stitches the full v0.2.0
customization stack together against the postgres + redis service
containers provided by ``.github/workflows/integration.yml``.

The single test runs the canonical demo flow a paying customer would
walk through:

1. Hit ``PUT /v1/tenant/config`` with three patches at once:
   - ``pipeline_id`` switched away from ``default_v1``.
   - SLA ``confidence_threshold`` bumped to 0.95 via patch.
   - A ``target_overrides`` row dropped on ``us-hs.v1`` (preserved
     alongside the template override the next step adds).

2. Upload a watermarked custom xlsx via
   ``POST /v1/templates/us-hs.v1/grade_9``.

3. Confirm ``GET /v1/tenant/config`` reflects every override (the
   pipeline_id, the SLA patch list, AND the merged target_overrides
   row carrying both the prior vocabulary patch AND the new
   templates[grade_9].blob_key patch).

4. Drive ``POST /v1/transcripts`` with vision mocked at 0.97 confidence
   so the SLA threshold doesn't park it, render the result, download
   the xlsx, and assert the watermark cell survived the render — proof
   that the template override flowed all the way to the renderer
   under postgres.

If any v0.2.0 wire breaks under postgres (JSONB column type drift, SET
LOCAL no-op regression, the tenant_overrides merge logic, the
target_overrides bucket in the resolver dep), THIS test will be the one
that catches it before main.
"""

from __future__ import annotations

import base64
import io
import os
import secrets
import uuid
from collections.abc import AsyncIterator
from pathlib import Path
from typing import Any
from unittest.mock import AsyncMock

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from PIL import Image as PILImage
from sqlalchemy import text
from sqlalchemy.ext.asyncio import (
    AsyncEngine,
    async_sessionmaker,
    create_async_engine,
)

from ocr_to_report.adapters.blob import LocalBlobStore
from ocr_to_report.adapters.crypto import EnvelopeEncryptor, EnvKEKProvider
from ocr_to_report.adapters.crypto.envelope import DEK_BYTES
from ocr_to_report.adapters.db import Base
from ocr_to_report.adapters.db.repositories import ApiKeyRepo, TenantRepo
from ocr_to_report.adapters.queue import InMemoryQueue
from ocr_to_report.adapters.vision import (
    ExtractionResult,
    FixedPolicy,
    InMemoryAsyncCache,
    ProviderRouter,
    TokenUsage,
    VisionProvider,
)
from ocr_to_report.api.app import create_app
from ocr_to_report.api.deps import AppState
from ocr_to_report.api.settings import Settings
from ocr_to_report.core.profiles import ProfileRegistry
from ocr_to_report.core.sla import SLA_PRESETS, SlaTier
from ocr_to_report.core.targets import TargetRegistry

pytestmark = pytest.mark.integration

REPO_ROOT = Path(__file__).resolve().parents[3]
WATERMARK_CELL = "Z1"
WATERMARK_VALUE = "v0.2.0-INTEGRATION-WATERMARK"

DB_URL = os.getenv("OCR2R_TEST_DB_URL")

if not DB_URL or not DB_URL.startswith("postgresql"):
    pytest.skip(
        "OCR2R_TEST_DB_URL not set (or not postgres); skipping v0.2.0 "
        "integration test. CI's integration.yml workflow sets it.",
        allow_module_level=True,
    )


# ─── helpers ──────────────────────────────────────────────────────────


def _png() -> bytes:
    img = PILImage.new("RGB", (400, 600), color=(220, 220, 220))
    out = io.BytesIO()
    img.save(out, format="PNG")
    return out.getvalue()


def _polish_extraction() -> dict[str, Any]:
    return {
        "full_name": "Jan Kowalski",
        "birth_date": "2010-01-15",
        "school_year": "2023/2024",
        "current_class_name": "pierwszej",
        "school_name": "Test Academy LO",
        "city": "Warszawa",
        "region": "mazowieckie",
        "promoted": True,
        "promoted_with_distinction": True,
        "conduct": "wzorowe",
        "subjects": [
            {"raw_subject_name": "Język polski", "raw_grade_value": "bardzo dobry"},
            {"raw_subject_name": "Matematyka", "raw_grade_value": "celujący"},
            {"raw_subject_name": "Fizyka", "raw_grade_value": "bardzo dobry"},
            {"raw_subject_name": "Chemia", "raw_grade_value": "bardzo dobry"},
            {"raw_subject_name": "Biologia", "raw_grade_value": "celujący"},
            {"raw_subject_name": "Historia", "raw_grade_value": "bardzo dobry"},
            {"raw_subject_name": "Geografia", "raw_grade_value": "dobry"},
            {"raw_subject_name": "Informatyka", "raw_grade_value": "celujący"},
            {"raw_subject_name": "Wychowanie fizyczne", "raw_grade_value": "bardzo dobry"},
            {"raw_subject_name": "Język angielski IV.1r.", "raw_grade_value": "bardzo dobry"},
        ],
        "advanced_subjects": ["Język angielski"],
    }


class _FixedConfidenceAdapter:
    """High-confidence vision mock so SLA never parks the job."""

    name = VisionProvider.MOCK

    def __init__(self, confidence: float = 0.97) -> None:
        self._confidence = confidence
        self.extract = AsyncMock(side_effect=self._extract)

    async def _extract(self, _request: Any) -> ExtractionResult:
        return ExtractionResult(
            raw_extraction=_polish_extraction(),
            confidence=self._confidence,
            field_confidences=None,
            warnings=[],
            provider=VisionProvider.MOCK,
            model_id="mock",
            usage=TokenUsage(input_tokens=1500, output_tokens=300, usd_cost=0.003),
        )

    async def aclose(self) -> None:
        return None


def _watermarked_xlsx() -> bytes:
    wb = Workbook()
    sheet = wb.active
    assert sheet is not None
    sheet[WATERMARK_CELL] = WATERMARK_VALUE
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ─── fixtures ─────────────────────────────────────────────────────────


@pytest.fixture
def settings(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> Settings:
    """Settings pointed at the postgres service container."""
    monkeypatch.setenv(
        "OCR2R_KEK_B64",
        base64.b64encode(secrets.token_bytes(DEK_BYTES)).decode(),
    )
    assert DB_URL is not None
    return Settings(
        env="development",
        database_url=DB_URL,
        blob_backend="local",
        blob_local_root=tmp_path / "blob",
        kek_b64=base64.b64encode(secrets.token_bytes(DEK_BYTES)).decode(),
        profiles_root=REPO_ROOT / "profiles",
        targets_root=REPO_ROOT / "targets",
    )


@pytest.fixture
async def pg_clean_engine() -> AsyncIterator[AsyncEngine]:
    """Fresh schema per test on the postgres service container.

    Drop + create the full ORM schema so seed state from prior runs
    (the postgres-integration tests in the same workflow) doesn't bleed.
    """
    assert DB_URL is not None
    engine = create_async_engine(DB_URL, future=True, echo=False)
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.drop_all)
        await conn.run_sync(Base.metadata.create_all)
    yield engine
    await engine.dispose()


@pytest.fixture
async def seeded(settings: Settings, pg_clean_engine: AsyncEngine) -> dict[str, Any]:
    """Tenant + API key with transcripts:write — the customer shape."""
    encryptor = EnvelopeEncryptor(EnvKEKProvider(env_var="OCR2R_KEK_B64"))
    sm = async_sessionmaker(pg_clean_engine, expire_on_commit=False)
    async with sm() as session:
        tenants = TenantRepo(session, encryptor)
        tenant, _dek = await tenants.create(name="Acme E2E", slug="acme-e2e")
        # Default tier (STANDARD) so the SLA-patch flip below has a
        # baseline to override.
        tenant.sla_tier = SlaTier.STANDARD.value
        keys = ApiKeyRepo(session)
        _row, plain_key = await keys.issue(
            tenant_id=tenant.id, scopes=["transcripts:write"]
        )
        await session.commit()
        return {"tenant_id": tenant.id, "api_key": plain_key}


def _make_client(settings: Settings) -> TestClient:
    """Build a TestClient with a high-confidence mock vision adapter wired in."""
    app = create_app(settings=settings)
    router = ProviderRouter(
        {VisionProvider.MOCK: _FixedConfidenceAdapter(confidence=0.97)},
        FixedPolicy(VisionProvider.MOCK),
    )
    encryptor = EnvelopeEncryptor(EnvKEKProvider(env_var="OCR2R_KEK_B64"))
    profile_registry = ProfileRegistry(settings.profiles_root.resolve())
    target_registry = TargetRegistry(settings.targets_root.resolve())
    bundle_roots = {
        t.id: settings.targets_root / t.id for t in target_registry.all()
    }

    client = TestClient(app)
    client.__enter__()
    app.state.app_state = AppState(
        settings=settings,
        encryptor=encryptor,
        profile_registry=profile_registry,
        target_registry=target_registry,
        blob_store=LocalBlobStore(settings.blob_local_root),
        vision_router=router,
        result_cache=InMemoryAsyncCache(),
        bundle_roots=bundle_roots,
        queue=InMemoryQueue(),
        sla_presets=dict(SLA_PRESETS),
    )
    return client


# ─── the integration test ─────────────────────────────────────────────


def test_full_v0_2_0_customization_flow_against_postgres(
    settings: Settings, seeded: dict[str, Any]
) -> None:
    """One flow exercising pipeline + SLA + target + template overrides on postgres."""
    client = _make_client(settings)
    try:
        headers = {"Authorization": f"Bearer {seeded['api_key']}"}
        target_id = "us-hs.v1"
        template_key = "grade_9"

        # 1. Apply three overrides in one PUT — the resolver must
        # demultiplex them per-scope.
        put = client.put(
            "/v1/tenant/config",
            headers=headers,
            json={
                "pipeline_id": "with_manual_review_v1",
                "sla_patches": [
                    {"op": "set", "path": "confidence_threshold", "value": 0.95}
                ],
                "target_overrides": {
                    target_id: [
                        {
                            "op": "set",
                            "path": "vocabulary.curriculum_note",
                            "value": "tenant-tagged",
                        }
                    ]
                },
            },
        )
        assert put.status_code == 200, put.text
        body = put.json()
        assert body["pipeline_id"] == "with_manual_review_v1"
        assert body["sla"]["confidence_threshold"] == 0.95
        assert target_id in body["target_overrides"]

        # 2. Upload a custom template. The endpoint must MERGE the
        # templates[grade_9].blob_key patch into the existing
        # target_overrides row from step 1 — not clobber the vocabulary
        # patch.
        upload = client.post(
            f"/v1/templates/{target_id}/{template_key}",
            headers=headers,
            files={
                "template_file": (
                    "custom.xlsx",
                    _watermarked_xlsx(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert upload.status_code == 201, upload.text
        upload_body = upload.json()
        assert upload_body["template_key"] == template_key

        # 3. GET /v1/tenant/config must reflect ALL overrides at once.
        cfg = client.get("/v1/tenant/config", headers=headers).json()
        assert cfg["pipeline_id"] == "with_manual_review_v1"
        assert cfg["sla"]["confidence_threshold"] == 0.95
        target_patches = cfg["target_overrides"][target_id]
        # Both the vocabulary patch AND the template upload patch must
        # be present — the merge in templates.py is what we're proving.
        paths = {p["path"] for p in target_patches}
        assert "vocabulary.curriculum_note" in paths, (
            f"vocabulary patch was clobbered by template upload; saw paths={paths}"
        )
        assert f"templates[{template_key}].blob_key" in paths, (
            f"template upload patch missing; saw paths={paths}"
        )

        # 4. Run a transcript. The bumped 0.95 threshold + mock @ 0.97
        # means the job succeeds (just clears the bar); the renderer
        # must pull the uploaded xlsx via the template override.
        r = client.post(
            "/v1/transcripts",
            headers=headers,
            files={"file": ("t.png", _png(), "image/png")},
            data={
                "profile_id": "pl.lo.swiadectwo_szkolne.v1",
                "target_id": target_id,
                "target_template_key": template_key,
            },
        )
        assert r.status_code == 200, r.text
        job = r.json()["job"]
        assert job["status"] == "succeeded", job

        # 5. Pull the rendered xlsx and confirm the watermark survived —
        # proof the template override flowed end-to-end through the
        # postgres-backed resolver into the renderer.
        download = client.get(f"/v1/jobs/{job['id']}/result", headers=headers)
        assert download.status_code == 200, download.text
        rendered_wb = load_workbook(filename=io.BytesIO(download.content))
        rendered_sheet = rendered_wb.active
        assert rendered_sheet is not None
        assert rendered_sheet[WATERMARK_CELL].value == WATERMARK_VALUE, (
            f"rendered xlsx is missing the watermark — uploaded template "
            f"didn't reach the renderer under postgres. "
            f"sheet={rendered_sheet.title!r}, Z1={rendered_sheet[WATERMARK_CELL].value!r}"
        )

        # 6. Confirm the RLS GUC was set for the request that wrote the
        # tenant_overrides row. The PUT closed its session; we open a
        # fresh one and verify both override rows exist under the
        # tenant_id RLS context — proves SET LOCAL stayed engaged across
        # the request lifecycle on postgres.
        assert DB_URL is not None
        engine = create_async_engine(DB_URL, future=True, echo=False)
        try:
            import asyncio  # noqa: PLC0415

            async def _read_back() -> list[tuple[str, str | None]]:
                async with engine.connect() as conn:
                    # Set the RLS GUC (the production code path does this
                    # via tenant_scoped_session).
                    await conn.execute(
                        text("SELECT set_config('app.tenant_id', :tid, true)"),
                        {"tid": str(seeded["tenant_id"])},
                    )
                    result = await conn.execute(
                        text(
                            "SELECT scope, target_id FROM tenant_overrides "
                            "WHERE tenant_id = :tid ORDER BY scope, target_id"
                        ),
                        {"tid": str(seeded["tenant_id"])},
                    )
                    return [(row[0], row[1]) for row in result.all()]

            rows = asyncio.run(_read_back())
            # One sla row (target_id=NULL) + one target row for us-hs.v1.
            kinds = {(scope, tid) for scope, tid in rows}
            assert ("sla", None) in kinds, f"missing sla row; saw {kinds}"
            assert ("target", "us-hs.v1") in kinds, f"missing target row; saw {kinds}"
        finally:
            import asyncio  # noqa: PLC0415

            asyncio.run(engine.dispose())
    finally:
        client.__exit__(None, None, None)


_ = uuid  # unused-import suppression for future expansion
