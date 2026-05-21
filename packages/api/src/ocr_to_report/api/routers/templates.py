"""``/v1/templates`` — list shipped templates + upload per-tenant overrides.

Three endpoints:

* ``GET  /v1/templates`` — list every target system + the template keys
  it ships. Read-only catalogue used by the web UI's "pick a target"
  dropdown and by SDK clients discovering what's available.

* ``POST /v1/templates/{target_id}/{template_key}`` — accept a multipart
  ``template_file`` upload (xlsx). The bytes are magic-byte-validated
  (PK\\x03\\x04 ZIP header) AND ``openpyxl.load_workbook``-validated
  (the renderer must actually be able to open it) before storage. The
  file lands in the blob store at
  ``tenant/{tenant.id}/templates/{target_id}/{template_key}/<sha256>.xlsx``
  and a ``tenant_overrides`` row is upserted with
  ``scope="template"``, ``target_id``, and a single ``set`` patch
  carrying the new ``blob_key``. From that point on every job the
  tenant runs against this target/template renders into the uploaded
  file's frame.

* ``DELETE /v1/templates/{target_id}/{template_key}`` — remove the
  override row (and the underlying blob) so jobs revert to the shipped
  template. The blob itself is also deleted to avoid orphan files
  accumulating in the per-tenant prefix; if multiple historical
  versions of the same key were uploaded, only the most-recent one is
  tracked by the override row, so older sha256-keyed blobs are
  unreachable and remain only as inert audit trail (operators can
  vacuum them out-of-band).
"""

from __future__ import annotations

import contextlib
import hashlib
from io import BytesIO
from typing import Annotated, Any

from fastapi import APIRouter, Depends, UploadFile, status
from openpyxl import load_workbook

from ocr_to_report.adapters.db.repositories import TenantOverrideRepo
from ocr_to_report.api.deps import AppState, RequestRepos, get_app_state, get_repos
from ocr_to_report.api.schemas import CustomTemplateResponse
from ocr_to_report.core.errors.domain import (
    NotFoundError,
    PayloadTooLargeError,
    UnsupportedMediaTypeError,
    ValidationError,
)

router = APIRouter(prefix="/v1", tags=["templates"])

# Magic header bytes for an xlsx file. xlsx is a ZIP archive, so any
# valid file starts with ``PK\x03\x04`` (the local file header sig).
# More-specific validation runs after this via openpyxl.
_XLSX_MAGIC = b"PK\x03\x04"

# Upper bound on a single template upload. xlsx templates are
# generally < 100 KB for transcript shapes; we cap an order of magnitude
# above that to catch accidental large-file uploads + give honest 413s.
_MAX_TEMPLATE_BYTES = 2_000_000

# Scope tag used on TenantOverride rows for uploaded templates.
#
# Custom-template uploads are persisted under the same ``target`` scope
# as other target-bundle patches (vocabulary edits, mapping tweaks). The
# patch shape — ``{"op": "set", "path": "templates[<key>].blob_key",
# "value": "<blob_key>"}`` — is what distinguishes a template upload
# inside that scope. This keeps the resolver dep (which buckets by
# scope) and the unique-key constraint ``(tenant, scope, target_id)``
# from forcing tenants into multiple rows for the same target_id.
_TEMPLATE_SCOPE = "target"


# ─── GET: catalogue of shipped templates ──────────────────────────────


@router.get("/templates")
async def list_templates(
    state: Annotated[AppState, Depends(get_app_state)],
) -> dict[str, Any]:
    """List every target system + the template keys it ships."""
    out = [
        {
            "target_id": target.id,
            "name": target.manifest.name,
            "version": target.manifest.version,
            "output_language": target.manifest.output_language,
            "output_formats": target.manifest.output_formats,
            "templates": [
                {
                    "key": t.key,
                    "output_format": t.output_format,
                    "target_year_index": t.target_year_index,
                }
                for t in target.templates
            ],
        }
        for target in state.target_registry.all()
    ]
    return {"targets": out}


# ─── POST: per-tenant custom template upload ──────────────────────────


@router.post(
    "/templates/{target_id}/{template_key}",
    response_model=CustomTemplateResponse,
    status_code=status.HTTP_201_CREATED,
    responses={
        400: {"description": "Upload is not a valid xlsx file"},
        401: {"description": "Authentication required"},
        404: {"description": "target_id or template_key unknown"},
        413: {"description": "Upload exceeds the size cap"},
        415: {"description": "Upload is not an xlsx (wrong magic bytes)"},
    },
)
async def upload_custom_template(
    target_id: str,
    template_key: str,
    state: Annotated[AppState, Depends(get_app_state)],
    repos: Annotated[RequestRepos, Depends(get_repos)],
    template_file: UploadFile,
) -> CustomTemplateResponse:
    """Persist a tenant-supplied xlsx as the active template for one slot.

    The upload's bytes replace the shipped template for THIS tenant on
    THIS ``(target_id, template_key)`` combination only. Cell bindings
    declared in the target bundle still drive what gets written — the
    upload changes the carrier (sheet layout, headers, frozen panes,
    styling), not the binding contract.
    """
    # 1. Validate the target/template exists in the shipped catalogue.
    target = state.target_registry.get(target_id)
    if target is None:
        raise NotFoundError(f"no target with id={target_id!r}", target_id=target_id)
    template = next((t for t in target.templates if t.key == template_key), None)
    if template is None:
        raise NotFoundError(
            f"target {target_id!r} has no template key {template_key!r}",
            target_id=target_id,
            template_key=template_key,
        )
    if template.output_format != "xlsx":
        # Uploading non-xlsx into an xlsx slot is the wrong endpoint shape.
        # When other formats are first-class (PDF templates, docx, …) this
        # endpoint should fork or take an output_format guard.
        raise ValidationError(
            f"upload override only supports xlsx templates; "
            f"{target_id!r}/{template_key!r} is {template.output_format!r}",
            target_id=target_id,
            template_key=template_key,
            output_format=template.output_format,
        )

    # 2. Read + size-check the upload.
    blob = await template_file.read()
    if len(blob) == 0:
        raise ValidationError("empty upload", size=0)
    if len(blob) > _MAX_TEMPLATE_BYTES:
        raise PayloadTooLargeError(
            f"template upload is {len(blob)} bytes; max is {_MAX_TEMPLATE_BYTES}",
            size=len(blob),
            max_bytes=_MAX_TEMPLATE_BYTES,
        )

    # 3. Magic-byte check. xlsx is a ZIP; the first 4 bytes must match.
    if not blob.startswith(_XLSX_MAGIC):
        raise UnsupportedMediaTypeError(
            "upload is not an xlsx file (missing PK\\x03\\x04 ZIP header)",
            size=len(blob),
        )

    # 4. openpyxl-level validation. Catches a ZIP that ISN'T an xlsx
    #    (e.g., a JAR, a custom ZIP) before we let it into the blob store.
    try:
        load_workbook(filename=BytesIO(blob))
    except Exception as e:
        raise ValidationError(
            f"openpyxl could not open the uploaded xlsx: {e}",
            template_key=template_key,
        ) from e

    # 5. Hash + store. The sha256 in the key makes repeated identical
    #    uploads idempotent at the blob layer; changing one cell yields
    #    a new key and a new override row revision.
    sha = hashlib.sha256(blob).hexdigest()
    blob_key = f"tenant/{repos.tenant.id}/templates/{target_id}/{template_key}/{sha}.xlsx"
    await state.blob_store.put(
        blob_key,
        blob,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    # 6. Merge into the override row. The row may already carry other
    #    target-scoped patches (vocabulary edits, mapping tweaks) for the
    #    same target_id; ``upsert`` replaces the entire patch list, so
    #    we read the existing row first, drop any prior blob_key patch
    #    for this template_key, and append the fresh one. The path
    #    syntax matches what the transcripts router scans for at render
    #    time: ``templates[<key>].blob_key``.
    overrides = TenantOverrideRepo(repos.session)
    existing_rows = await overrides.list_for_tenant(
        repos.tenant.id, scope=_TEMPLATE_SCOPE, include_disabled=True
    )
    existing_row = next((r for r in existing_rows if r.target_id == target_id), None)
    merged_patches: list[dict[str, Any]] = []
    if existing_row is not None:
        merged_patches = [
            p for p in existing_row.patches if not _is_blob_key_set_for(p, template_key)
        ]
    merged_patches.append(
        {
            "op": "set",
            "path": f"templates[{template_key}].blob_key",
            "value": blob_key,
        }
    )
    await overrides.upsert(
        tenant_id=repos.tenant.id,
        scope=_TEMPLATE_SCOPE,
        target_id=target_id,
        patches=merged_patches,
    )
    await repos.session.commit()

    return CustomTemplateResponse(
        target_id=target_id,
        template_key=template_key,
        blob_key=blob_key,
        sha256=sha,
        size_bytes=len(blob),
    )


# ─── DELETE: revert to the shipped template ───────────────────────────


@router.delete(
    "/templates/{target_id}/{template_key}",
    status_code=status.HTTP_204_NO_CONTENT,
    responses={
        401: {"description": "Authentication required"},
        404: {"description": "No custom template exists for this slot"},
    },
)
async def delete_custom_template(
    target_id: str,
    template_key: str,
    state: Annotated[AppState, Depends(get_app_state)],
    repos: Annotated[RequestRepos, Depends(get_repos)],
) -> None:
    """Remove the per-tenant template override; revert to the shipped template.

    Both the override row and the underlying blob are removed. If the
    blob delete fails (already gone, network blip), the row removal
    still wins — the worst case is an orphaned blob, which is a storage
    cleanup concern not a correctness one.
    """
    overrides = TenantOverrideRepo(repos.session)
    rows = await overrides.list_for_tenant(
        repos.tenant.id,
        scope=_TEMPLATE_SCOPE,
        include_disabled=True,
    )
    row = next((r for r in rows if r.target_id == target_id), None)
    if row is None:
        raise NotFoundError(
            f"no custom template uploaded for {target_id!r}/{template_key!r}",
            target_id=target_id,
            template_key=template_key,
        )

    # Find the blob_key in the row's patch list. Multiple template keys
    # share one row per target_id, so scan rather than assume index 0.
    # When the target_id row exists but carries no patch for THIS
    # template_key (e.g. only ``grade_9`` is overridden but the caller
    # is DELETE-ing ``grade_10``), we have to 404 explicitly — without
    # this guard the handler falls through to a no-op upsert that leaves
    # the row unchanged but returns 204, lying to the caller about
    # whether anything actually got deleted.
    blob_key = _blob_key_from_patches(row.patches, template_key)
    if blob_key is None:
        raise NotFoundError(
            f"no custom template uploaded for {target_id!r}/{template_key!r}",
            target_id=target_id,
            template_key=template_key,
        )

    # Mutate the row's patch list to drop just this template_key's entry.
    remaining = [p for p in row.patches if not _is_blob_key_set_for(p, template_key)]
    if not remaining:
        await overrides.delete(
            tenant_id=repos.tenant.id,
            scope=_TEMPLATE_SCOPE,
            target_id=target_id,
        )
    else:
        await overrides.upsert(
            tenant_id=repos.tenant.id,
            scope=_TEMPLATE_SCOPE,
            target_id=target_id,
            patches=remaining,
        )
    await repos.session.commit()

    # Best-effort blob cleanup. Failure here doesn't change the override
    # state — the next render will fall through to the shipped file. The
    # broad ``Exception`` swallow is deliberate: blob store ops can fail
    # for unrelated reasons (network blip, race with another delete),
    # and a 500 here would hide the successful row removal.
    if blob_key is not None:
        with contextlib.suppress(Exception):
            await state.blob_store.delete(blob_key)


# ─── helpers ──────────────────────────────────────────────────────────


def _is_blob_key_set_for(patch: dict[str, Any], template_key: str) -> bool:
    """Match a patch like ``{"op": "set", "path": "templates[<k>].blob_key", ...}``."""
    if not isinstance(patch, dict):
        return False
    return patch.get("op") == "set" and patch.get("path") == f"templates[{template_key}].blob_key"


def _blob_key_from_patches(patches: list[dict[str, Any]], template_key: str) -> str | None:
    """Pull the ``blob_key`` value out of a patch list for one template key."""
    for p in patches:
        if _is_blob_key_set_for(p, template_key):
            v = p.get("value")
            if isinstance(v, str):
                return v
    return None


__all__ = ["router"]
