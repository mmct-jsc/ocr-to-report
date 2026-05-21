"""openpyxl-based Excel renderer.

Loads the target bundle's xlsx template, writes each cell binding's
resolved value into the corresponding cell, and returns the file's
bytes. Template formatting (fonts, borders, colors, merged cells) is
preserved — we only mutate cell **values**.

Cells whose binding resolved to ``None`` are left untouched, so any
literal "-" or pre-filled hint in the template stays in place.
"""

from __future__ import annotations

import io
from pathlib import Path
from typing import TYPE_CHECKING

from openpyxl import load_workbook

from ocr_to_report.adapters.render.protocol import RendererError

if TYPE_CHECKING:
    from ocr_to_report.core.mapping import RenderData
    from ocr_to_report.core.target.bundle import TargetBundle


class XlsxRenderer:
    """Stateless openpyxl-based renderer.

    Construct with the bundle root (absolute filesystem path); call
    instances as ``renderer(target_bundle, render_data) -> bytes``.
    """

    def __init__(self, bundle_root: Path) -> None:
        if not bundle_root.is_dir():
            raise RendererError(
                f"bundle root not a directory: {bundle_root}",
                bundle_root=str(bundle_root),
            )
        self._root = bundle_root

    def __call__(
        self,
        target_bundle: TargetBundle,
        render_data: RenderData,
        *,
        template_override_bytes: bytes | None = None,
    ) -> bytes:
        return render_xlsx(
            target_bundle,
            render_data,
            bundle_root=str(self._root),
            template_override_bytes=template_override_bytes,
        )


def render_xlsx(
    target_bundle: TargetBundle,
    render_data: RenderData,
    *,
    bundle_root: str,
    template_override_bytes: bytes | None = None,
) -> bytes:
    """Fill the chosen template's xlsx with the render data and return bytes.

    When ``template_override_bytes`` is provided, the renderer treats that
    payload as the template's xlsx file — the on-disk
    ``bundle_root / template.blob_path`` is bypassed. This is how
    per-tenant custom-template uploads (v0.2.0 Task 6) route through:
    the API endpoint stores the uploaded xlsx in the blob store, persists
    a ``templates[<key>].blob_key`` override patch, and the transcripts
    router fetches those bytes back and passes them here. The bindings
    in the target bundle still drive which cells get written — the
    upload changes the *carrier* (frame, styling, header), not the
    binding contract.
    """
    template = next(
        (t for t in target_bundle.templates if t.key == render_data.template_key),
        None,
    )
    if template is None:
        raise RendererError(
            f"target bundle {target_bundle.id!r} has no template with key "
            f"{render_data.template_key!r}",
            target_id=target_bundle.id,
            template_key=render_data.template_key,
        )
    if template.output_format != "xlsx":
        raise RendererError(
            f"render_xlsx called with non-xlsx template (format={template.output_format!r})",
        )

    if template_override_bytes is not None:
        size = len(template_override_bytes)
        try:
            workbook = load_workbook(filename=io.BytesIO(template_override_bytes))
        except Exception as e:
            raise RendererError(
                f"could not load tenant-uploaded template ({size} bytes): {e}",
                template_key=render_data.template_key,
                override_bytes=size,
            ) from e
    else:
        template_path = Path(bundle_root) / template.blob_path
        if not template_path.is_file():
            raise RendererError(
                f"template file missing: {template_path}",
                template_path=str(template_path),
            )

        try:
            workbook = load_workbook(filename=template_path)
        except Exception as e:
            raise RendererError(
                f"could not load template {template_path}: {e}",
                template_path=str(template_path),
            ) from e

    sheet = workbook.active
    if sheet is None:  # pragma: no cover — openpyxl returns the active sheet
        raise RendererError("template workbook has no active sheet")

    for cell_ref, value in render_data.cells.items():
        if value is None:
            # Don't overwrite a pre-filled "-" or hint already in the template.
            continue
        try:
            sheet[cell_ref] = value
        except Exception as e:
            raise RendererError(
                f"failed writing cell {cell_ref}: {e}",
                cell_ref=cell_ref,
            ) from e

    out = io.BytesIO()
    workbook.save(out)
    return out.getvalue()


__all__ = ["XlsxRenderer", "render_xlsx"]
